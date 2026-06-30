"""
data_io.py — Importación y exportación de datos
=====================================================
Toda la lógica de:
  - Leer un CSV/Excel subido por el usuario y convertirlo en filas
    que ml_models.py pueda consumir
  - Exportar resultados de clustering a un Excel formateado
    y a un PDF de reporte ejecutivo

Separado de app.py (rutas) y de ml_models.py (algoritmos) para que
cada pieza tenga una sola responsabilidad.
"""

import io
import pandas as pd
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm
from reportlab.lib import colors as rl_colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


# Columnas requeridas para que un registro pueda pasar por ambos modelos
REQUIRED_COLUMNS = [
    'INDICE_OCUPACIONAL', 'PROMEDIO_DIAS_ESTADA', 'INDICE_ROTACION',
    'NUMERO_EGRESOS', 'PROMEDIO_CAMAS_DISPONIBLE', 'LETALIDAD', 'MES',
]


# ════════════════════════════════════════════════════════════
#  IMPORTACIÓN
# ════════════════════════════════════════════════════════════
def read_uploaded_table(file_storage) -> pd.DataFrame:
    """
    Lee un archivo subido (CSV o Excel) y devuelve un DataFrame crudo.
    file_storage: objeto de Flask request.files['archivo']
    """
    filename = file_storage.filename.lower()

    if filename.endswith('.csv'):
        raw_bytes = file_storage.read()
        # Intenta separador ; (formato chileno/REM20) y si falla, ,
        try:
            df = pd.read_csv(io.BytesIO(raw_bytes), sep=';')
            if df.shape[1] == 1:  # no se separó bien, probablemente es coma
                df = pd.read_csv(io.BytesIO(raw_bytes), sep=',')
        except Exception:
            df = pd.read_csv(io.BytesIO(raw_bytes), sep=',')
    elif filename.endswith(('.xlsx', '.xls')):
        df = pd.read_excel(file_storage)
    else:
        raise ValueError('Formato no soportado. Usa .csv, .xlsx o .xls')

    df.columns = [str(c).strip().upper() for c in df.columns]
    return df


def validate_columns(df: pd.DataFrame) -> dict:
    """Verifica que el archivo tenga las columnas mínimas necesarias."""
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    return {
        'valid': len(missing) == 0,
        'missing_columns': missing,
        'found_columns': list(df.columns),
        'row_count': len(df),
    }


def dataframe_to_rows(df: pd.DataFrame) -> list[dict]:
    """Convierte el DataFrame importado en lista de dicts para ml_models."""
    df = df.fillna(0)
    return df.to_dict(orient='records')


# ════════════════════════════════════════════════════════════
#  EXPORTACIÓN — EXCEL
# ════════════════════════════════════════════════════════════
HEADER_FILL  = PatternFill('solid', start_color='0D3B3E', end_color='0D3B3E')
HEADER_FONT  = Font(color='FFFFFF', bold=True, size=10, name='Arial')
TITLE_FONT   = Font(color='0D3B3E', bold=True, size=16, name='Arial')
SUB_FONT     = Font(color='5C6B6C', size=10, italic=True, name='Arial')
BORDER_THIN  = Border(*(Side(style='thin', color='E8E3D8') for _ in range(4)))

CLUSTER_FILLS = {
    'C0': PatternFill('solid', start_color='FFF7ED', end_color='FFF7ED'),
    'C1': PatternFill('solid', start_color='ECFDF3', end_color='ECFDF3'),
    'C2': PatternFill('solid', start_color='EFF6FF', end_color='EFF6FF'),
    'C3': PatternFill('solid', start_color='FEF2F2', end_color='FEF2F2'),
}


def build_excel_report(df_results: pd.DataFrame, km_meta: dict) -> bytes:
    """
    Construye un Excel de 2 hojas:
      1. Resumen ejecutivo (métricas agregadas + gráficos)
      2. Detalle de clasificaciones (tabla completa)
    Devuelve los bytes del archivo, listos para enviar como descarga.
    """
    wb = Workbook()

    ws = wb.active
    ws.title = 'Resumen Ejecutivo'

    ws['B2'] = 'REM 20 — Reporte de Patrones Operativos'
    ws['B2'].font = TITLE_FONT
    ws['B3'] = f"Generado el {datetime.now().strftime('%d-%m-%Y %H:%M')} · {len(df_results)} registros analizados"
    ws['B3'].font = SUB_FONT
    ws.merge_cells('B2:F2')
    ws.merge_cells('B3:F3')

    valid = df_results[df_results['_error'].isna()] if '_error' in df_results.columns else df_results

    kpi_row = 6
    kpis = [
        ('Registros Procesados', len(df_results)),
        ('Patrones Identificados', valid['CLUSTER'].nunique() if 'CLUSTER' in valid.columns else 'N/A'),
        ('Patrón Más Frecuente', f"C{valid['CLUSTER'].mode().iloc[0]}" if 'CLUSTER' in valid.columns and len(valid) else 'N/A'),
        ('Errores', int(df_results['_error'].notna().sum()) if '_error' in df_results.columns else 0),
    ]
    for i, (label, value) in enumerate(kpis):
        col = get_column_letter(2 + i)
        ws[f'{col}{kpi_row}'] = label
        ws[f'{col}{kpi_row}'].font = Font(bold=True, size=9, color='5C6B6C', name='Arial')
        ws[f'{col}{kpi_row+1}'] = value
        ws[f'{col}{kpi_row+1}'].font = Font(bold=True, size=18, color='0D3B3E', name='Arial')

    chart_row = kpi_row + 4
    if 'CLUSTER' in valid.columns and len(valid):
        ws[f'B{chart_row}'] = 'Distribución por Patrón Operativo (Cluster)'
        ws[f'B{chart_row}'].font = Font(bold=True, size=11, color='0D3B3E', name='Arial')

        cluster_counts = valid['CLUSTER'].value_counts().sort_index()
        cl_data_start = chart_row + 1
        ws[f'B{cl_data_start}'] = 'Cluster'
        ws[f'C{cl_data_start}'] = 'Patrón'
        ws[f'D{cl_data_start}'] = 'Cantidad'
        for c in 'BCD':
            ws[f'{c}{cl_data_start}'].font = HEADER_FONT
            ws[f'{c}{cl_data_start}'].fill = HEADER_FILL

        for i, (cluster_id, count) in enumerate(cluster_counts.items()):
            r = cl_data_start + 1 + i
            name = km_meta['cluster_names'].get(str(cluster_id), f'Cluster {cluster_id}')
            ws[f'B{r}'] = f"C{cluster_id}"
            ws[f'C{r}'] = name
            ws[f'D{r}'] = int(count)

        chart = BarChart()
        chart.title = 'Registros por Patrón Operativo'
        chart.y_axis.title = 'Cantidad'
        chart.x_axis.title = 'Patrón'
        chart.style = 10
        n = len(cluster_counts)
        data_ref = Reference(ws, min_col=4, min_row=cl_data_start, max_row=cl_data_start + n)
        cats_ref = Reference(ws, min_col=2, min_row=cl_data_start + 1, max_row=cl_data_start + n)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width, chart.height = 16, 9
        ws.add_chart(chart, f'F{chart_row}')

    for col, width in zip('ABCDEF', [3, 12, 32, 14, 4, 4]):
        ws.column_dimensions[col].width = width

    ws2 = wb.create_sheet('Detalle de Clasificaciones')

    export_cols = [c for c in df_results.columns if not c.startswith('_')]
    priority = ['CLUSTER', 'PATRON_OPERATIVO']
    ordered = [c for c in priority if c in export_cols] + [c for c in export_cols if c not in priority]

    for j, col_name in enumerate(ordered, start=1):
        cell = ws2.cell(row=1, column=j, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    for i, (_, row) in enumerate(df_results.iterrows(), start=2):
        for j, col_name in enumerate(ordered, start=1):
            val = row.get(col_name, '')
            cell = ws2.cell(row=i, column=j, value=val)
            cell.border = BORDER_THIN
            if col_name == 'PATRON_OPERATIVO' and isinstance(val, str):
                key = f"C{row.get('CLUSTER', '')}"
                if key in CLUSTER_FILLS:
                    cell.fill = CLUSTER_FILLS[key]

    for j, col_name in enumerate(ordered, start=1):
        ws2.column_dimensions[get_column_letter(j)].width = max(14, len(str(col_name)) + 2)
    ws2.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════
#  EXPORTACIÓN — PDF
# ════════════════════════════════════════════════════════════
def build_pdf_report(df_results: pd.DataFrame, km_meta: dict) -> bytes:
    """
    Construye un PDF de reporte ejecutivo: portada con KPIs,
    distribución de patrones operativos, y tabla resumen
    de los registros de los clusters menos frecuentes.
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        topMargin=2 * cm, bottomMargin=2 * cm,
        leftMargin=2 * cm, rightMargin=2 * cm,
    )
    styles = getSampleStyleSheet()

    ink = rl_colors.HexColor('#0D3B3E')
    teal = rl_colors.HexColor('#1A6B6F')
    steel = rl_colors.HexColor('#5C6B6C')

    title_style = ParagraphStyle('TitleClin', parent=styles['Title'], textColor=ink, fontSize=20, spaceAfter=4)
    sub_style = ParagraphStyle('SubClin', parent=styles['Normal'], textColor=steel, fontSize=10, fontName='Helvetica-Oblique')
    h2_style = ParagraphStyle('H2Clin', parent=styles['Heading2'], textColor=ink, fontSize=13, spaceBefore=14, spaceAfter=8)
    body_style = ParagraphStyle('BodyClin', parent=styles['Normal'], fontSize=9.5, leading=14)

    story = []
    story.append(Paragraph('Reporte de Patrones Operativos', title_style))
    story.append(Paragraph('Sistema REM 20 — Clustering K-Means de registros hospitalarios', sub_style))
    story.append(Paragraph(f"Generado el {datetime.now().strftime('%d-%m-%Y a las %H:%M')}", sub_style))
    story.append(Spacer(1, 16))

    valid = df_results[df_results['_error'].isna()] if '_error' in df_results.columns else df_results

    story.append(Paragraph('Resumen General', h2_style))
    kpi_data = [['Métrica', 'Valor']]
    kpi_data.append(['Registros procesados', str(len(df_results))])
    kpi_data.append(['Registros clasificados OK', str(len(valid))])
    if 'CLUSTER' in valid.columns and len(valid):
        kpi_data.append(['Patrones distintos detectados', str(valid['CLUSTER'].nunique())])
        for cid in sorted(valid['CLUSTER'].unique()):
            name = km_meta['cluster_names'].get(str(cid), f'Cluster {cid}')
            count = int((valid['CLUSTER'] == cid).sum())
            pct = count / len(valid) * 100
            kpi_data.append([f"  C{cid} — {name}", f"{count}  ({pct:.1f}%)"])

    t = Table(kpi_data, colWidths=[11 * cm, 4 * cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), ink),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#FAFAF7')]),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E8E3D8')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(t)

    if 'CLUSTER' in valid.columns and len(valid):
        story.append(Paragraph('Distribución de Patrones Operativos', h2_style))
        cluster_counts = valid['CLUSTER'].value_counts().sort_index()
        cl_data = [['Cluster', 'Patrón', 'Registros', '% del Total']]
        for cid, count in cluster_counts.items():
            name = km_meta['cluster_names'].get(str(cid), f'Cluster {cid}')
            pct = count / len(valid) * 100
            cl_data.append([f"C{cid}", name, str(int(count)), f"{pct:.1f}%"])

        t2 = Table(cl_data, colWidths=[1.8 * cm, 7.5 * cm, 2.7 * cm, 3 * cm])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), teal),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#FAFAF7')]),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E8E3D8')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(t2)

        if 'INDICE_OCUPACIONAL' in valid.columns:
            story.append(Paragraph('Indicadores Promedio por Patrón', h2_style))
            agg_cols = [c for c in ['INDICE_OCUPACIONAL', 'LETALIDAD', 'PROMEDIO_DIAS_ESTADA', 'INDICE_ROTACION', 'NUMERO_EGRESOS'] if c in valid.columns]
            agg = valid.groupby('CLUSTER')[agg_cols].mean().round(2)
            head = ['Patrón'] + agg_cols
            rows = [head]
            for cid, r in agg.iterrows():
                name = km_meta['cluster_names'].get(str(cid), f'Cluster {cid}')
                rows.append([f"C{cid} — {name}"] + [str(r[c]) for c in agg_cols])
            col_w = [4.5 * cm] + [2.3 * cm] * len(agg_cols)
            t4 = Table(rows, colWidths=col_w, repeatRows=1)
            t4.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), teal),
                ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8.5),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#FAFAF7')]),
                ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#E8E3D8')),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(t4)

    story.append(Spacer(1, 20))
    story.append(Paragraph(
        'Reporte generado automáticamente por el sistema REM 20. '
        'Modelo: K-Means (K=4, clustering de patrones operativos). '
        'Proyecto de Minería de Datos — BIY7121.',
        ParagraphStyle('Footer', parent=body_style, fontSize=7.5, textColor=steel),
    ))

    doc.build(story)
    buf.seek(0)
    return buf.read()
